from openpyxl import load_workbook
from config import *
import ast
import time

def duplicate(data):
    workbook = load_workbook(filename=config['spreadsheet_path'])
    sheet = workbook.active
    i=1
    while True:
        if sheet['A' + str(i)].value == None:
            return False
        if sheet['A' + str(i)].value == data:
            print(f'DUPLICATE DATA - {data}')
            return True
        i=i+1


def convert(orginal_path, converted_path):
    file = open(orginal_path, 'r')
    converted_file = open(converted_path, 'w+')
    lines = file.readlines()
    for line in lines:
        info = line.split('; ')
        print(info)
        for inf in info:
            try:
                print(inf.split(' ')[1])
                inf = inf.replace('\n', '')
                fname = inf.split(' ')[0]
                email = inf.split(' <')[1].replace('<', '').replace('>', '').replace(';', '')
                lname = inf.replace(fname + ' ', '').replace(email, '').replace(' <>', '').replace(';', '')
                # print(f'\nFirst name - {fname}\nLast name - {lname}\nEmail - {email}\n')
            except:
                email = inf.split(' ')[0].replace(';','')
                fname=''
                lname=''
                print('Website')
            array = [email, fname.replace(',',''), lname]
            converted_file.write(f'{str(array)}\n')
    file.close()
    converted_file.close()


def spreadsheet_check(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    print(workbook.sheetnames)
    i=1
    while True:
        if sheet['A'+str(i)].value == None:
            number = i
            break
        i=i+1
    print(number)
    return number


def spreadsheet_write(number, file_path, spreadsheet_path):
    from openpyxl import Workbook
    conv_file = open(file_path)
    lines = conv_file.readlines()
    workbook = load_workbook(filename=spreadsheet_path)
    sheet = workbook.active
    temp=0
    duplicate_values = 0
    for i in range(number,number+len(lines)):
        line = lines[temp]
        line = ast.literal_eval(line)
        if not duplicate(str(line[0])):
            ie=i-duplicate_values
            sheet['A'+str(ie)] = str(line[0])
            sheet['B'+str(ie)] = str(line[1])
            sheet['C'+str(ie)] = str(line[2])
        else:
            duplicate_values = duplicate_values + 1
            print(f'Duplicate Detected- {line[0]}, Ignoring')
        temp=temp+1
    workbook.save(filename=spreadsheet_path)
    conv_file.close()

def main():
    convert(config['list_path'], config['converted_file_path'])
    spreadsheet_write(spreadsheet_check(config['spreadsheet_path']), config['converted_file_path'], config['spreadsheet_path'])
try:
    main()
except Exception as e:
    print(e)
time.sleep(3)
